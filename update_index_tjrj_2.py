import re
import pandas as pd
import pdfplumber
from datetime import datetime
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def extract_correction_factors(pdf_path, debug=False):
    """Extract monthly correction factors from the PDF."""
    correction_factors = {}
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_matches = 0
            
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                
                # Define regex pattern to match the format in the PDF (01/MM/YYYY followed by a number)
                pattern = r'01/(\d{2})/(\d{4})\s+(\d+,\d+)'
                
                # Find all matches
                matches = re.findall(pattern, text)
                total_matches += len(matches)
                
                for month_str, year_str, factor_str in matches:
                    # Parse month and year
                    month = int(month_str)
                    year = int(year_str)
                    
                    # Convert factor string to float (replace comma with period)
                    factor = float(factor_str.replace(',', '.'))
                    
                    # Store in dictionary
                    correction_factors[(month, year)] = factor
                
                # Print debug info if requested
                if debug and page_num < 2:
                    print(f"\nPage {page_num+1} sample: {text[:200]}...")
                    print(f"Found {len(matches)} matches on page {page_num+1}")
            
            print(f"Total matches found across all pages: {total_matches}")
    except Exception as e:
        print(f"Error extracting from PDF: {e}")
        return {}
    
    return correction_factors

def find_column_indices(sheet, date_column_name, rate_column_name):
    """Find the column indices for date and rate columns."""
    # Get the header row
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    date_col_idx = None
    rate_col_idx = None
    
    # Find the column indices
    for idx, cell_value in enumerate(header_row, 1):
        if cell_value == date_column_name:
            date_col_idx = idx
        if cell_value == rate_column_name:
            rate_col_idx = idx
    
    return date_col_idx, rate_col_idx

def update_excel_with_openpyxl(excel_path, correction_factors, output_path, date_column, rate_column, debug=False):
    """Update the Excel spreadsheet with correction factors while preserving formatting and formulas."""
    try:
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Assume the first sheet if it's not specified
        sheet = wb.active
        
        # Find column indices
        date_col_idx, rate_col_idx = find_column_indices(sheet, date_column, rate_column)
        
        if not date_col_idx:
            print(f"Error: Column '{date_column}' not found in the Excel file")
            if debug:
                print(f"Available columns in the first row: {[cell.value for cell in sheet[1]]}")
            return None
        
        # If rate column doesn't exist, create it
        if not rate_col_idx:
            # Find the first empty column
            max_col = sheet.max_column
            rate_col_idx = max_col + 1
            # Add header
            sheet.cell(row=1, column=rate_col_idx, value=rate_column)
            print(f"Created new column '{rate_column}' at position {rate_col_idx}")
        
        # Process each row
        updated_count = 0
        not_found_count = 0
        
        # Start from row 2 (assuming row 1 is header)
        for row_idx in range(2, sheet.max_row + 1):
            date_cell = sheet.cell(row=row_idx, column=date_col_idx)
            
            if date_cell.value:
                # Convert to datetime if it's not already
                date_value = date_cell.value
                if isinstance(date_value, str):
                    try:
                        date_value = datetime.strptime(date_value, '%d/%m/%Y')
                    except ValueError:
                        try:
                            # Try with other common formats
                            date_value = pd.to_datetime(date_value).to_pydatetime()
                        except:
                            print(f"Could not parse date: {date_value} in row {row_idx}")
                            continue
                
                # Extract month and year
                if hasattr(date_value, 'month') and hasattr(date_value, 'year'):
                    month = date_value.month
                    year = date_value.year
                    
                    # Look up correction factor
                    key = (month, year)
                    if key in correction_factors:
                        # Update the cell while preserving formatting
                        sheet.cell(row=row_idx, column=rate_col_idx, value=correction_factors[key])
                        updated_count += 1
                    else:
                        not_found_count += 1
                        if debug:
                            print(f"No factor found for: {month}/{year}")
                else:
                    if debug:
                        print(f"Row {row_idx} has date value but it's not a proper date: {date_value}")
            else:
                if debug and row_idx < 10:  # Only show for first few rows to avoid spam
                    print(f"Row {row_idx} has no date value")
        
        print(f"Updated {updated_count} rows, could not find factors for {not_found_count} rows")
        
        # Save updated Excel
        wb.save(output_path)
        
        return output_path
    except Exception as e:
        print(f"Error updating Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def main(pdf_path, excel_path, output_path, date_column, rate_column, debug=False):
    """Main function to orchestrate the extraction and update process."""
    # Check if files exist
    if not os.path.exists(pdf_path):
        print(f"Error: PDF file not found at {pdf_path}")
        return None
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return None
    
    # Extract correction factors from PDF
    print("Extracting correction factors from PDF...")
    correction_factors = extract_correction_factors(pdf_path, debug)
    print(f"Extracted {len(correction_factors)} correction factors")
    
    if not correction_factors:
        print("No correction factors extracted. Check the PDF format.")
        return None
    
    # Sample some factors for verification
    if debug:
        print("\nSample correction factors:")
        sample_keys = list(correction_factors.keys())[:5]
        for key in sample_keys:
            print(f"Month: {key[0]}, Year: {key[1]} -> Factor: {correction_factors[key]}")
    
    # Update Excel with correction factors
    print("\nUpdating Excel file...")
    result_path = update_excel_with_openpyxl(excel_path, correction_factors, output_path, date_column, rate_column, debug)
    
    if result_path:
        print(f"Done! Updated Excel saved to {result_path}")
    else:
        print("Failed to update Excel file.")
    
    return result_path

if __name__ == "__main__":
    # Parse command-line arguments or use defaults
    if len(sys.argv) >= 6:
        pdf_path = sys.argv[1]
        excel_path = sys.argv[2]
        output_path = sys.argv[3]
        date_column = sys.argv[4]
        rate_column = sys.argv[5]
        debug = len(sys.argv) > 6 and sys.argv[6].lower() == 'debug'
    else:
        # Default values - adjust as needed
        pdf_path = "Relatório de Correção Monetária.pdf"
        excel_path = "04 Planilha Débito  G II S201 26.02.2025 1.xls"
        output_path = "updated_spreadsheet.xlsx"
        date_column = "Data"  # Replace with actual date column name
        rate_column = "Taxa"  # Replace with desired rate column name
        debug = True
        
        print("Using default values. To specify values, use:")
        print("python extract_rates.py pdf_path excel_path output_path date_column rate_column [debug]")
    
    main(pdf_path, excel_path, output_path, date_column, rate_column, debug)