import re
import pandas as pd
import pdfplumber
from datetime import datetime
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def extract_correction_factors(pdf_path, debug=False):
    """Extract monthly correction factors from the PDF."""
    correction_factors = {}
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_matches = 0
            
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                
                # Define regex pattern to match the format in the PDF (01/MM/YYYY followed by a number)
                pattern = r'01/(\d{2})/(\d{4})\s+(\d+,\d+)'
                
                # Find all matches
                matches = re.findall(pattern, text)
                total_matches += len(matches)
                
                for month_str, year_str, factor_str in matches:
                    # Parse month and year
                    month = int(month_str)
                    year = int(year_str)
                    
                    # Convert factor string to float (replace comma with period)
                    factor = float(factor_str.replace(',', '.'))
                    
                    # Store in dictionary
                    correction_factors[(month, year)] = factor
                
                # Print debug info if requested
                if debug and page_num < 2:
                    print(f"\nPage {page_num+1} sample: {text[:200]}...")
                    print(f"Found {len(matches)} matches on page {page_num+1}")
            
            print(f"Total matches found across all pages: {total_matches}")
    except Exception as e:
        print(f"Error extracting from PDF: {e}")
        return {}
    
    return correction_factors

def find_column_indices(sheet, column_name):
    """Find all occurrences of a column name in the sheet."""
    indices = []
    
    # Iterate through all cells in row 1
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col_idx).value
        if cell_value == column_name:
            indices.append(col_idx)
    
    return indices

def update_excel_with_openpyxl(excel_path, correction_factors, output_path, sheet_mappings, debug=False):
    """Update the Excel spreadsheet with correction factors while preserving formatting and formulas.
    
    Args:
        sheet_mappings: List of tuples (sheet_name, date_column, rate_column) for each table to update
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_path)
        
        total_updated_count = 0
        total_not_found_count = 0
        
        # Process each sheet mapping
        for mapping_idx, (sheet_name, date_column, rate_column) in enumerate(sheet_mappings):
            print(f"\nProcessing sheet: {sheet_name}, Columns: {date_column} -> {rate_column}")
            
            # Get the sheet by name or index
            if sheet_name.isdigit():
                # If sheet_name is a number, get by index (0-based, add 1 for user-friendly numbering)
                sheet_idx = int(sheet_name) - 1
                if sheet_idx < 0 or sheet_idx >= len(wb.sheetnames):
                    print(f"Warning: Sheet index {sheet_name} is out of range. Available sheets: {wb.sheetnames}")
                    continue
                sheet = wb.worksheets[sheet_idx]
                print(f"  Using sheet at index {sheet_name}: '{sheet.title}'")
            else:
                # Get by name
                if sheet_name not in wb.sheetnames:
                    print(f"Warning: Sheet '{sheet_name}' not found in the workbook. Available sheets: {wb.sheetnames}")
                    continue
                sheet = wb[sheet_name]
            
            # Find all occurrences of the date column
            date_col_indices = find_column_indices(sheet, date_column)
            
            if not date_col_indices:
                print(f"Warning: Column '{date_column}' not found in sheet '{sheet.title}'")
                continue
            
            # For each date column, find the corresponding rate column
            for date_col_idx in date_col_indices:
                # Find rate column
                # First try by name
                rate_col_indices = find_column_indices(sheet, rate_column)
                
                # If we found rate columns, use the one that's closest to the date column
                if rate_col_indices:
                    # Find the rate column closest to this date column
                    rate_col_idx = min(rate_col_indices, key=lambda x: abs(x - date_col_idx))
                else:
                    # If not found by name, look for headers in the same row group as the date column
                    # This assumes the rate column is within a few columns of the date column
                    found = False
                    # Look up to 5 columns to the right
                    for col_offset in range(1, 6):
                        check_col = date_col_idx + col_offset
                        if check_col <= sheet.max_column:
                            cell_value = sheet.cell(row=1, column=check_col).value
                            if cell_value == rate_column:
                                rate_col_idx = check_col
                                found = True
                                break
                    
                    if not found:
                        print(f"Warning: Could not find '{rate_column}' column near '{date_column}' column at position {date_col_idx}")
                        continue
                
                print(f"  Found table with date column at {date_col_idx} and rate column at {rate_col_idx}")
                
                # Process each row
                updated_count = 0
                not_found_count = 0
                
                # Start from row 2 (assuming row 1 is header)
                for row_idx in range(2, sheet.max_row + 1):
                    date_cell = sheet.cell(row=row_idx, column=date_col_idx)
                    
                    if date_cell.value:
                        # Convert to datetime if it's not already
                        date_value = date_cell.value
                        if isinstance(date_value, str):
                            try:
                                date_value = datetime.strptime(date_value, '%d/%m/%Y')
                            except ValueError:
                                try:
                                    # Try with other common formats
                                    date_value = pd.to_datetime(date_value).to_pydatetime()
                                except:
                                    if debug:
                                        print(f"  Could not parse date: {date_value} in row {row_idx}")
                                    continue
                        
                        # Extract month and year
                        if hasattr(date_value, 'month') and hasattr(date_value, 'year'):
                            month = date_value.month
                            year = date_value.year
                            
                            # Look up correction factor
                            key = (month, year)
                            if key in correction_factors:
                                # Update the cell while preserving formatting
                                sheet.cell(row=row_idx, column=rate_col_idx, value=correction_factors[key])
                                updated_count += 1
                            else:
                                not_found_count += 1
                                if debug:
                                    print(f"  No factor found for: {month}/{year}")
                        else:
                            if debug:
                                print(f"  Row {row_idx} has date value but it's not a proper date: {date_value}")
                
                print(f"  Updated {updated_count} rows, could not find factors for {not_found_count} rows")
                total_updated_count += updated_count
                total_not_found_count += not_found_count
        
        print(f"\nTotal updates: {total_updated_count} rows updated, {total_not_found_count} rows without matching factors")
        
        # Save updated Excel
        wb.save(output_path)
        
        return output_path
    except Exception as e:
        print(f"Error updating Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def main(pdf_path, excel_path, output_path, sheet_mappings, debug=False):
    """Main function to orchestrate the extraction and update process."""
    # Check if files exist
    if not os.path.exists(pdf_path):
        print(f"Error: PDF file not found at {pdf_path}")
        return None
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return None
    
    # Extract correction factors from PDF
    print("Extracting correction factors from PDF...")
    correction_factors = extract_correction_factors(pdf_path, debug)
    print(f"Extracted {len(correction_factors)} correction factors")
    
    if not correction_factors:
        print("No correction factors extracted. Check the PDF format.")
        return None
    
    # Sample some factors for verification
    if debug:
        print("\nSample correction factors:")
        sample_keys = list(correction_factors.keys())[:5]
        for key in sample_keys:
            print(f"Month: {key[0]}, Year: {key[1]} -> Factor: {correction_factors[key]}")
    
    # Update Excel with correction factors
    print("\nUpdating Excel file...")
    result_path = update_excel_with_openpyxl(excel_path, correction_factors, output_path, sheet_mappings, debug)
    
    if result_path:
        print(f"Done! Updated Excel saved to {result_path}")
    else:
        print("Failed to update Excel file.")
    
    return result_path

def print_usage():
    """Print script usage instructions."""
    print("\nUsage:")
    print("python extract_rates.py pdf_path excel_path output_path sheet1 date_col1 rate_col1 [sheet2 date_col2 rate_col2] [debug]")
    print("\nParameters:")
    print("  pdf_path: Path to the PDF containing correction factors")
    print("  excel_path: Path to the Excel spreadsheet")
    print("  output_path: Where to save the updated spreadsheet")
    print("  sheet1: Sheet name or number (1-based) for the primary table")
    print("  date_col1: Name of the date column in the primary table")
    print("  rate_col1: Name of the column where rates should be inserted in the primary table")
    print("  sheet2: (Optional) Sheet name or number for the secondary table")
    print("  date_col2: (Optional) Name of the date column in the secondary table")
    print("  rate_col2: (Optional) Name of the column where rates should be inserted in the secondary table")
    print("  debug: (Optional) Add 'debug' as the last parameter to show detailed information")
    print("\nExample:")
    print('python extract_rates.py "report.pdf" "data.xlsx" "updated.xlsx" 1 "Data" "Taxa" 2 "Data" "Fator Corr." debug')

if __name__ == "__main__":
    # Parse command-line arguments
    arg_count = len(sys.argv)
    debug = "debug" in sys.argv[-1].lower()
    
    # Define default sheet mappings
    sheet_mappings = []
    
    if arg_count >= 7:
        pdf_path = sys.argv[1]
        excel_path = sys.argv[2]
        output_path = sys.argv[3]
        primary_sheet = sys.argv[4]
        primary_date_col = sys.argv[5]
        primary_rate_col = sys.argv[6]
        
        # Add primary sheet mapping
        sheet_mappings.append((primary_sheet, primary_date_col, primary_rate_col))
        
        # Check if secondary sheet mapping is provided
        if arg_count >= 10:
            secondary_sheet = sys.argv[7]
            secondary_date_col = sys.argv[8]
            secondary_rate_col = sys.argv[9]
            sheet_mappings.append((secondary_sheet, secondary_date_col, secondary_rate_col))
        
        # Run the main function
        main(pdf_path, excel_path, output_path, sheet_mappings, debug)
    else:
        # Default values for manual testing
        pdf_path = "Relatório de Correção Monetária.pdf"
        excel_path = "04 Planilha Débito  G II S201 26.02.2025 1.xls"
        output_path = "updated_spreadsheet.xlsx"
        sheet_mappings = [
            ("1", "Data", "Taxa"),              # Sheet 1, primary table
            ("2", "Data", "Fator Corr.")        # Sheet 2, secondary table
        ]
        
        print("Not enough arguments provided. Using default values for testing.")
        print(f"PDF: {pdf_path}")
        print(f"Excel: {excel_path}")
        print(f"Output: {output_path}")
        print("Sheet mappings:")
        for idx, mapping in enumerate(sheet_mappings):
            print(f"  Mapping {idx+1}: Sheet '{mapping[0]}', Date column = '{mapping[1]}', Rate column = '{mapping[2]}'")
        
        # Print usage instructions
        print_usage()
        
        # Ask for confirmation before running with defaults
        response = input("\nDo you want to continue with these defaults? (y/n): ")
        if response.lower() == 'y':
            main(pdf_path, excel_path, output_path, sheet_mappings, debug=True)
        else:
            print("Exiting. Please run the script with the required arguments.")